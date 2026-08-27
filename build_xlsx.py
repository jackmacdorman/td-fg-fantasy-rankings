#!/usr/bin/env python3
"""Build the draft board workbook from data/board.json.

Points/VOR are written as live formulas pointing at the Settings tab, so editing
a projection or a scoring weight re-ranks the sheet instead of silently
disagreeing with the numbers next to it.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
BOARD = json.loads((ROOT / "data" / "board.json").read_text())
PLAYERS = BOARD["players"]
REPL = BOARD["replacement"]

FONT = "Arial"
POS_ORDER = ["QB", "RB", "WR", "TE", "K", "DST"]
RELIABILITY = {"QB": 0.65, "RB": 0.60, "WR": 0.55, "TE": 0.55, "K": 0.35, "DST": 0.20}

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT, color="0000FF", size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True, color="1F3864")

POS_FILL = {
    "QB": "FCE4D6", "RB": "E2EFDA", "WR": "DDEBF7",
    "TE": "FFF2CC", "K": "EDEDED", "DST": "E4DFEC",
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)

COLS = [
    ("Rank", 6), ("Pos", 6), ("Player", 24), ("Tm", 5), ("P", 5),
    ("PassTD", 8), ("RushTD", 8), ("RecTD", 8), ("FG", 6), ("PAT", 6),
    ("DefTD", 7), ("Saf", 6),
    ("Points", 9), ("VOR", 8), ("Adj VOR", 9), ("Conf", 8), ("Notes", 62),
]
# Column letters for the stat inputs, used to build the Points formula.
F_PASS, F_RUSH, F_REC, F_FG, F_PAT, F_DEF, F_SAF = "FGHIJKL"


def style_header(ws, row=1):
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sheet(ws, rows, title):
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws.row_dimensions[1].height = 20

    for i, (name, _) in enumerate(COLS, 1):
        ws.cell(row=2, column=i, value=name)
    style_header(ws, row=2)
    ws.freeze_panes = "C3"

    for r, p in enumerate(rows, start=3):
        vals = [
            p["overall_rank"], p["pos_rank"], p["name"], p["team"], p["pos"],
            p["pass_td"] or None, p["rush_td"] or None, p["rec_td"] or None,
            p["fg"] or None, p["pat"] or None,
            p["dst_td"] or None, p["safety"] or None,
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = INPUT_FONT if 6 <= i <= 12 else BODY
            cell.border = BORDER
            if i in (1, 2, 4, 5) or 6 <= i <= 12:
                cell.alignment = Alignment(horizontal="center")

        pts = (
            f"={F_PASS}{r}*Settings!$B$4+{F_RUSH}{r}*Settings!$B$5"
            f"+{F_REC}{r}*Settings!$B$6+{F_FG}{r}*Settings!$B$7"
            f"+{F_PAT}{r}*Settings!$B$8+{F_DEF}{r}*Settings!$B$9"
            f"+{F_SAF}{r}*Settings!$B$10"
        )
        repl = f"INDEX(Settings!$G$4:$G$9,MATCH($E{r},Settings!$F$4:$F$9,0))"
        rely = f"INDEX(Settings!$H$4:$H$9,MATCH($E{r},Settings!$F$4:$F$9,0))"

        ws.cell(row=r, column=13, value=pts).font = BOLD
        ws.cell(row=r, column=14, value=f"=M{r}-{repl}").font = BODY
        ws.cell(row=r, column=15, value=f"=N{r}*{rely}").font = BOLD
        ws.cell(row=r, column=16, value=p.get("confidence", "")).font = BODY
        ws.cell(row=r, column=17, value=p.get("notes", "")).font = BODY

        for i in (13, 14, 15):
            c = ws.cell(row=r, column=i)
            c.number_format = "0.0"
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
        ws.cell(row=r, column=16).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=16).border = BORDER
        ws.cell(row=r, column=17).border = BORDER

        fill = PatternFill("solid", fgColor=POS_FILL[p["pos"]])
        for i in (2, 5):
            ws.cell(row=r, column=i).fill = fill

    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{len(rows) + 2}"


def settings_sheet(ws):
    ws["A1"] = "League Settings — edit the blue cells and every tab re-ranks"
    ws["A1"].font = TITLE
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    for col in "DEFGH":
        ws.column_dimensions[col].width = 14

    ws["A3"] = "SCORING"
    ws["A3"].font = Font(name=FONT, bold=True, size=11)
    scoring = [
        ("Passing TD", 3), ("Rushing TD", 6), ("Receiving TD", 6),
        ("Field Goal (any distance)", 3), ("Extra Point", 1),
        ("Def/ST TD", 6), ("Safety", 2),
    ]
    for i, (label, val) in enumerate(scoring, start=4):
        ws.cell(row=i, column=1, value=label).font = BODY
        c = ws.cell(row=i, column=2, value=val)
        c.font = INPUT_FONT
        c.alignment = Alignment(horizontal="center")

    ws["A12"] = "Teams in league"
    ws["A12"].font = BODY
    ws["B12"] = BOARD["teams"]
    ws["B12"].font = INPUT_FONT
    ws["B12"].alignment = Alignment(horizontal="center")

    ws["F3"] = "Pos"
    ws["G3"] = "Replacement"
    ws["H3"] = "Reliability"
    for col in "FGH":
        c = ws[f"{col}3"]
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, pos in enumerate(POS_ORDER, start=4):
        ws.cell(row=i, column=6, value=pos).font = BODY
        rc = ws.cell(row=i, column=7, value=round(REPL[pos], 2))
        rc.font = INPUT_FONT
        rc.number_format = "0.0"
        lc = ws.cell(row=i, column=8, value=RELIABILITY[pos])
        lc.font = INPUT_FONT
        lc.number_format = "0.00"
        for col in (6, 7, 8):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")

    notes = [
        "",
        "Replacement level = mean points of the players just past the last startable slot",
        "at each position (12 teams x starters). Computed by rank.py; recomputed if you",
        "change the player pool. Starters: 1QB, 2RB, 2WR, 1TE, 1K, 1D/ST.",
        "",
        "Reliability = how much of a projected edge actually shows up in reality. These are",
        "JUDGMENT ESTIMATES, not fitted from data. Kicker FG totals and defensive/return TDs",
        "have very little year-over-year predictive power, so most of their apparent spread",
        "is noise; raw VOR badly overrates both. Set every value to 1.00 to see raw VOR.",
        "",
        "Blue cells are inputs. Black cells are formulas. Edit projections directly on any",
        "tab and Points/VOR/Adj VOR recalculate.",
    ]
    for i, line in enumerate(notes, start=11):
        ws.cell(row=i, column=6, value=line).font = Font(name=FONT, size=9, italic=True)


def simple_sheet(ws, title, headers, rows, widths):
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[2].height = 24
    for r, row in enumerate(rows, start=3):
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY
            c.alignment = Alignment(vertical="top", wrap_text=(widths[i - 1] > 30))
            c.border = BORDER
    ws.freeze_panes = "A3"


SLEEPERS = [
    ["Sean Tucker", "TB", "RB", "Owns the Tampa goal line. Bucky Irving had ZERO carries inside the 5 in 2025; Tucker took nearly all of them and scored 8 TDs, 6 in his final 8 games. TB re-signed him specifically for that role. His ~2.8 projected TDs is a model artifact of low carry volume. Biggest value in the format."],
    ["Jaxson Dart", "NYG", "QB", "Grades QB8 on ~6 rushing TDs (9 as a rookie, 2nd-highest designed-rush share among QBs). Rushing TDs are worth 2x passing TDs here. Caveat: new HC John Harbaugh spent a decade teaching Lamar Jackson to slide — highest-variance pick on the board."],
    ["Daniel Jones", "IND", "QB", "Clay projects 6 rushing TDs, Razzball 7.3. Grades QB17 in this format. Nobody drafts him for that."],
    ["Ka'imi Fairbairn", "HOU", "K", "Purest FG-lean in the league (~50% of value from FGs, not PATs). Houston led the NFL with 52 FG attempts in 2025 because the offense stalls in the red zone — exactly what you want when a FG is worth 3x a PAT."],
    ["Daniel Carlson", "NO", "K", "Signed Aug 24, 2026 — every published projection was built on Charlie Smyth and is stale. New Orleans took 42 FG attempts in 2025 (T-3rd) at a league-worst 71.4%. An accurate kicker on that volume is unpriced."],
    ["Chris Rodriguez Jr.", "JAX", "RB", "Hand-picked by Liam Coen as the goal-line back over Bhayshul Tuten. Pure TD specialist, free in drafts."],
    ["Justin Fields", "KC", "QB", "Best handcuff in the format: a rushing QB sitting behind Mahomes, who is coming off ACL+LCL surgery and has not played a preseason snap."],
    ["Malik Willis", "MIA", "QB", "Pure lottery ticket. Yahoo projects 7 rushing TDs (would be 4th among QBs), Razzball 2.8. If Miami uses him on designed goal-line runs he is a top-15 QB from an undrafted price."],
]

TRAPS = [
    ["Bucky Irving", "TB", "RB", "Zero carries inside the 5-yard line in 2025. Costs a real pick and does not get the goal line. Sean Tucker does."],
    ["Saquon Barkley", "PHI", "RB", "Jalen Hurts vultures ~8.5 rushing TDs on the tush push, which survived the 2026 rules vote. Barkley's ADP does not price this in. (James Cook has the same Josh Allen problem, but his 10.5 Vegas line already accounts for it.)"],
    ["Bijan Robinson", "ATL", "RB", "Took only 48% of Atlanta's inside-5 carries in 2025, and the team signed Brian Robinson Jr. A committee back at a bell-cow price."],
    ["Tyler Bass", "BUF", "K", "Buffalo attempted a league-low 21 field goals in 2025 — they score touchdowns instead. Second-worst FG lean in the league. Fine in standard formats, bad here."],
    ["Eagles D/ST", "PHI", "DST", "Will be top-5 in every conventional D/ST ranking, which is built on sacks and points allowed — both worth ZERO in this league. No return threat and only 5 defensive TDs in three years. Ranks 25th here."],
    ["Matthew Stafford", "LAR", "QB", "34 passing TDs but 0.5 rushing. Elite in standard leagues, mid-pack here because passing TDs are only worth 3."],
    ["Jared Goff", "DET", "QB", "Same problem: 30 passing TDs, no rushing floor."],
    ["Kyren Williams / Blake Corum", "LAR", "RB", "Literally alternate drives, and reporting says Davante Adams is the Rams' actual goal-line target — note Adams projects for ~10 receiving TDs."],
]

STRATEGY = [
    ["1", "Rushing QBs are the single biggest edge", "A rushing TD (6) is worth double a passing TD (3). Josh Allen projects for 153 points — 30 clear of QB2 and a bigger gap than QB2-to-QB12. He is a legitimate 1.01 in this format. After him, prioritize rushing floor over passing volume every time."],
    ["2", "Draft your kicker around round 2", "Counterintuitive but real. Kickers have the highest raw VOR of any position after Allen, because FG x 3 produces huge absolute totals. BUT kicker projections are the least reliable of any position, so the risk-adjusted board slots K1 near pick 17 rather than pick 4. That is still four to six rounds earlier than anyone else in your league will move."],
    ["3", "Target offenses that stall, not offenses that score", "For kickers only. A drive ending in a FG pays 3; a drive ending in a TD pays the kicker 1. Houston (52 FGA) is the ideal kicker offense. Buffalo (21 FGA) is the worst."],
    ["4", "Chase the goal-line role, not the workload", "Yardage is worth nothing. A short-yardage specialist who scores 10 times beats a 1,400-yard back who scores 5. Identify who gets the ball inside the 5 in every backfield — that is the whole RB position."],
    ["5", "Stream D/ST, do not draft one early", "The D1-to-D32 spread is only 11 points, versus 66 for kickers. With no sacks or points-allowed scoring, defenses only score on return/defensive TDs — a stat that barely correlates year to year. Take one in the last round. If your roster allows, a second kicker is worth more than a second defense."],
    ["6", "Punt returners matter more than pass rushers", "Punt return TDs jumped from 6 to 15 league-wide in 2025 while kick return TDs fell. When you do pick a defense, buy the return game (SEA, DAL, NE, NYJ, TEN) over the conventional 'elite defense' (PHI)."],
    ["7", "Fade elite possession receivers", "Target volume is worth nothing without the end zone. Prioritize contested-catch and red-zone bodies. Note Davante Adams (LAR) projects for ~10 receiving TDs as the Rams' goal-line target — WR2 overall here despite a modest standard-league ADP."],
]

SOURCES = [
    ["Vegas — RB rushing TD O/U", "Nine real season-long lines: Henry 12.5, Taylor 11.5, Cook 10.5, Kyren Williams 9.5, Javonte Williams 9.5, Montgomery 7.5, Chase Brown 5.5, Achane 5.5, Mason 4.5. Books are LOWER than consensus on Chase Brown and Mason, HIGHER on Henry."],
    ["Vegas — RB receiving TD O/U", "Three lines only: Gibbs O4.5, McCaffrey O4.5, Bijan O3.5."],
    ["Vegas — QB passing TD O/U", "Four lines only: Allen O24.5, Prescott O27.5, Herbert O22.5, Stroud U22.5."],
    ["Vegas — QB rushing TD", "NONE EXIST at any book. Confirmed. Since rushing TDs carry most of the weight for mobile QBs, this is the single largest uncertainty on the board."],
    ["Vegas — WR/TE and kickers", "No usable season-long TD props found. Kicker futures are not posted by major books at all; only weekly in-season FG totals."],
    ["Kalshi prediction markets", "Season rushing-TD ladders were pulled but DELIBERATELY EXCLUDED — zero volume, 20-40 cent spreads, and non-monotonic pricing (Henry's 12+ traded above his 10+). Used only as a sanity check on ordering. Raw data in data/raw/rb_sources/."],
    ["QB projections", "ESPN/Mike Clay 2026 Projection Guide (updated 8/26/26, all 32 team pages parsed), Razzball, Yahoo, FFToday, Footballguys, CBS, FantasyPros."],
    ["RB projections", "FFToday full 2026 table (94 RBs, rush/rec TD split) + ESPN projections API (102 RBs) as independent second source, plus Ourlads depth charts for goal-line roles."],
    ["WR/TE projections", "FFToday 2026 tables + ESPN projections API, merged with name/team normalization. 81 of 117 players confirmed by both sources; the rest are single-source and flagged LOW confidence."],
    ["Kicker projections", "CBS Sports 2026 season projections (FGM/XPM split for all 32 starters), cross-checked against FantasyPros top 10 and Draft Sharks. Supplemented with 2025 team FG attempts from NFL.com as the red-zone-stalling proxy."],
    ["D/ST projections", "MODEL ESTIMATE, not sourced. Built from NFL.com three-year defensive TD history (2023-25) regressed 50% to the league mean of 1.6, plus a returner-quality adjustment and a flat blocked-kick term. Model totals 74 TDs vs. ~72 actual in 2025."],
    ["Known gaps", "No 2026 red zone TD% table obtained (all sources 403/404). No projection published for A.J. Dillon (CAR), a plausible goal-line sleeper. Nick Chubb could not be found on any 2026 roster — omitted rather than guessed at."],
    ["Reliability caveat", "Confidence flags mean 2+ sources agree within ~2 TDs. For deep backups that is a weak signal — two models agreeing a backup scores ~1 TD is not real information. Treat HIGH as meaningful only in the top ~45 at RB and top ~30 at WR."],
]


def main():
    wb = Workbook()

    ws = wb.active
    ws.title = "Draft Board"
    write_sheet(ws, PLAYERS, "2026 TD/FG-Only Draft Board — ranked by Adjusted VOR")

    for pos in POS_ORDER:
        rows = sorted(
            (p for p in PLAYERS if p["pos"] == pos), key=lambda p: -p["points"]
        )
        label = {"DST": "DEF"}.get(pos, pos)  # "/" is illegal in a sheet name
        write_sheet(wb.create_sheet(label), rows, f"{label} — ranked by projected points")

    simple_sheet(
        wb.create_sheet("Strategy"), "How this format differs from normal fantasy",
        ["#", "Principle", "Why"],
        STRATEGY, [4, 38, 105],
    )
    simple_sheet(
        wb.create_sheet("Sleepers"), "Players the projections systematically underrate",
        ["Player", "Tm", "Pos", "Why he is undervalued in a TD-only league"],
        SLEEPERS, [22, 6, 6, 110],
    )
    simple_sheet(
        wb.create_sheet("Traps"), "Players who cost more than they are worth here",
        ["Player", "Tm", "Pos", "Why to fade"],
        TRAPS, [26, 6, 6, 110],
    )
    simple_sheet(
        wb.create_sheet("Sources"), "Where every number came from, and what is missing",
        ["Category", "Detail"],
        SOURCES, [30, 115],
    )
    settings_sheet(wb.create_sheet("Settings"))

    out = ROOT / "2026 TD-FG League Draft Board.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
